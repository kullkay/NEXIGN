from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pandas import DataFrame

from json import load


input_file = "portainer.json"
output_file = "portainer.xlsx"


with open(input_file, "r") as f:
    data = load(f)


vulns = []
for result in data.get("Results", []):
    for vuln in result.get("Vulnerabilities", []):
        vulns.append({
            "PkgName": vuln.get("PkgName"),
            "VulnerabilityID": vuln.get("VulnerabilityID"),  
            "PkgID": vuln.get("PkgID"),
            "Severity": vuln.get("Severity"),
        })

df = DataFrame(vulns)

grouped_df = DataFrame(vulns).sort_values(by=["PkgName", "VulnerabilityID"])


grouped_df.to_excel(output_file, index=False)

wb = load_workbook(output_file)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

wb.save(output_file)